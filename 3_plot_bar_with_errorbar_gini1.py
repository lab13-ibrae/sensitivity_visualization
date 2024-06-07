# -*- coding: utf-8 -*-
import numpy as np
import os
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.legend_handler import HandlerPatch
from matplotlib.patches import FancyArrowPatch
import pandas as pd
from openpyxl import load_workbook


#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'

filename='snl_no_graph_gini_bootstrap_qoi1.xlsx'

filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    
#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"

#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]


# #dataframe with stats
sens_data_qoi_stat=sens_data_qoi.describe().iloc[1:3,1:].T.reset_index().rename(columns={"index": "parameter","mean": "importance","std": "importance_std"})

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")
fig, ax = plt.subplots(1,1,figsize=(8,6))

#prepare chart elements
#lables by parameter names
labels = sens_data_qoi_stat['parameter']

#positions of the bars
x_pos=np.arange(len(labels))

bar=ax.bar(x_pos, sens_data_qoi_stat['importance'],
            yerr=sens_data_qoi_stat['importance_std'], capsize=5, 
            #align='center', alpha=0.5, color='dimgray',ecolor='0.4',hatch='')
            align='center', alpha=0.5, color='0.2',ecolor='0.3',hatch='')
ax.set_ylabel('Importance measures')
ax.set_xlabel('Parameters')
ax.set_xticks(x_pos)
ax.set_xticklabels(labels)
bar_patch = mpatches.Rectangle((0, 0), 1, 2,color='0.2',alpha=0.5, label='Mean importance')
err_patch =FancyArrowPatch((0, 0), (1, 1), color='0.3',label='Standard deviation')


def make_err_legend(legend, orig_handle,
                      xdescent, ydescent,
                      width, height, fontsize):
    p = mpatches.FancyArrowPatch((0, 3), (20, 3), mutation_scale=5, arrowstyle='|-|' )
    return p


ax.legend(handles=[bar_patch, err_patch],handler_map={mpatches.FancyArrowPatch : HandlerPatch(patch_func=make_err_legend)})



ax.yaxis.grid(True)
plt.title(f'Gini importance measure. Statistics via bootstrap over 50 train/test splits.\nQoI: {qoi}', y=1, fontweight="bold")



# Save the figure and show
plt.tight_layout()

plt.savefig(f'3_{filename_prefix}_{qoi}_bar_with_err.png', dpi=600)