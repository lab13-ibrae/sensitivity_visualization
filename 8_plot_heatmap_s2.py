# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import pickle
import seaborn as sns
import matplotlib.pyplot as plt
import networkx as nx
import itertools
import matplotlib.collections as mcoll
import matplotlib.path as mpath
from matplotlib.colors import LinearSegmentedColormap
import colorsys
import matplotlib.colors as mpl_colors
from openpyxl import load_workbook

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
filename='snl_no_graph_s2_dakota_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = filename)
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
   
#qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]

#add missing cells for full symmetric matrix
all_param_names=list(dict.fromkeys(sens_data_qoi['parameter1'].unique().tolist() + sens_data_qoi['parameter2'].unique().tolist()))
for par in all_param_names:
    self_dummy={'S2':0,
                'parameter1':par,
                'parameter2':par,
                    }
    sens_data_qoi=pd.concat([sens_data_qoi,pd.DataFrame([self_dummy])])
s2_df=sens_data_qoi.pivot_table(index='parameter1',columns='parameter2',values='S2',sort=False)
s2_df=s2_df[all_param_names]
s2_df=s2_df.fillna(s2_df.T).fillna(0)

#mask
triang = np.triu(s2_df)
triang[np.triu_indices_from(triang)]=1

#colormap
#custom_cmap = sns.diverging_palette(220, 130, s=75, l=50, center='dark',as_cmap=True)
custom_cmap = sns.diverging_palette(220, 20, as_cmap=True)
fig, ax = plt.subplots(figsize=(8, 8))
g = sns.heatmap(s2_df,cmap=custom_cmap,  alpha=1,linewidths=.9,linecolor='1',vmin=0,
                annot=True, fmt=".3f",
                mask=triang,
                cbar_kws={'location':"right","shrink": 0.5,"orientation": "vertical","pad": -0.15
                          }
                )
#ax.xaxis.set_ticks_position("top")
ax.tick_params(left=False, top=False,bottom=False,pad=-2)
ax.set_xticks(ax.get_xticks()[:-1])
ax.set_yticks(ax.get_yticks()[1:])
ax.set(xlabel=None)
ax.set(ylabel=None)
cbar = ax.collections[0].colorbar
cbar.set_label('$S_2$ index',fontsize=12,labelpad=0, y=1.1,x=-0.9, ha='right', rotation=0)


#fig.tight_layout()
fig.subplots_adjust(top=1.01,bottom=0.15,right=1.01,)
fig.suptitle(f"Sobol' 2nd order indices (interactions).\nQoI: {qoi}",y=0.98,fontweight="bold",fontsize=13)

fig.savefig(f"8_{filename_prefix}_heatmap_s2.png", dpi=600)
