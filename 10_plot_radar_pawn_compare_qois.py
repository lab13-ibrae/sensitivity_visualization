###################
# radar chart

# -*- coding: utf-8 -*-
#import required libraries
import os
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from math import pi

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'
filename='snl_no_graph_pawn_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    



#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")
##~Okabe&Ito palette
palette= ['#F0E442',          '#E69F00',
          '#D55E00',          '#0072B2',
          '#009E73',          '#56B4E9',
          '#CC79A7',          '000000',]
#create figure, set dimensions
fig, ax = plt.subplots(1,1,figsize=(8,6),subplot_kw=dict(projection="polar"))

# Create background
#count parameters
parameters=list(sens_data[qoi]['parameter'])
num_par=len(parameters)
 
# calculate angles of each axis in the plot (divide the plot by number of parameters)
angles = [n / float(num_par) * 2 * pi for n in range(num_par)]
angles += angles[:1]
 
 
# position of the first axis :
ax.set_theta_offset(pi / 2)
ax.set_theta_direction(-1)
 
# Draw one axe per parameters + add labels

 
# Draw ylabels
ax.set_rlabel_position(0)
plt.ylim(0,0.6)
ticks=[0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55]
ticklabels=[f"{tick}" for tick in ticks]
plt.yticks(ticks, ticklabels, color="grey", size=7,zorder=100)
 
# Plot each line = each compared set of the data


#qoi1
qoi='Peak_TotalI129_M'
qoi_color=palette[3]
si_values=list(sens_data[qoi]['mean'].values)
si_values += si_values[:1]

ax.plot(angles, si_values, linewidth=1, linestyle='solid', color=qoi_color,label=qoi)
ax.fill(angles, si_values, qoi_color, alpha=0.1)

#qoi2
qoi='FractionofSpikeinRepository_1My'
qoi_color=palette[4]
si_values=list(sens_data[qoi]['mean'].values)
si_values += si_values[:1]

ax.plot(angles, si_values, linewidth=1, linestyle='solid', color=qoi_color,label=qoi)
ax.fill(angles, si_values, qoi_color, alpha=0.1)


#qoi3
qoi='AqEb_RockEb_1Myr'
qoi_color=palette[6]
si_values=list(sens_data[qoi]['mean'].values)
si_values += si_values[:1]

ax.plot(angles, si_values, linewidth=1, linestyle='solid', color=qoi_color,label=qoi)
ax.fill(angles, si_values, qoi_color, alpha=0.1)


plt.xticks(angles[:-1], parameters)
ax.tick_params(axis='x',pad=15)
# Add legend
plt.legend(loc='upper right', bbox_to_anchor=(0.35, 0.01), fontsize=12)
fig.subplots_adjust(top=0.85,bottom=0.15,left=0.15,right=0.85)
fig.suptitle(f"PAWN sensitivity index by QoI",fontsize=13,fontweight='bold',y=0.98)


#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables                                    
plt.savefig(f'10_{filename_prefix}_radar_pawn_by_qoi.png',dpi=600)