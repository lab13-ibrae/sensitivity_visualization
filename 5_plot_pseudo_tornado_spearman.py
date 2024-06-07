# -*- coding: utf-8 -*-
import os
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.cm
import matplotlib.colors

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'
filename='snl_no_graph_spearman_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    
#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]


#sort by sensitivity indice absolute value
sens_data_qoi=sens_data_qoi.reindex(sens_data_qoi['SRCC'].abs().sort_values(ascending=False).index)

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid",rc={'patch.force_edgecolor':True,
                   'patch.edgecolor': 'lightgray'})
fig, ax = plt.subplots(1,1,figsize=(8,6))

#create figure, set dimensions



#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }
#map palette to the parameters
sens_data_qoi['color']=sens_data_qoi['parameter'].map(palette)

def cmap_exists(name):
    try:
         matplotlib.cm.get_cmap(name)
         return True
    except ValueError:
         pass
    return False

#register custom colors as matplotlib colormap
len_colors=len(list(sens_data_qoi['color']))
custom_colormap=LinearSegmentedColormap.from_list('custom_snl_colors',
                                                  list(sens_data_qoi['color']),
                                                  N=len_colors)
if not cmap_exists('custom_snl_colors'):
    matplotlib.colormaps.register(custom_colormap)
    
custom_pal = sns.color_palette("custom_snl_colors", n_colors=len_colors, 
                               #desat=0.2
                               )

#barplot variant form seaborn library
#plot colored total indices
g = sns.barplot(ax=ax,data=sens_data_qoi, y="parameter", x='SRCC',palette=custom_pal,)



#customize labels
g.tick_params(labelsize=11)
plt.ylabel('')
plt.xlabel('')

#set limits for X-axis
ax.set_xlim([-0.4, 0.4])

ax.yaxis.grid(True) 
#plot title
plt.title(f" Spearman rank correlation coefficients. QoI: {qoi}", y=1, fontweight="bold")

#adjust the padding
plt.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables
plt.savefig(f'5_{filename_prefix}_{qoi}_pseudo_tornado_spearman.png', dpi=600)



