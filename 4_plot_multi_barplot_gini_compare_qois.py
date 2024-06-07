# -*- coding: utf-8 -*-
#import required libraries
import os
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from matplotlib.lines import Line2D
#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into one dataframe
data_folder='./sensitivity_results_datasets'
filename='snl_no_graph_gini_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")
wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data=pd.DataFrame()
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    qoi_tmp=pd.DataFrame(data, columns=cols)
    qoi_tmp['qoi']=qoi
    sens_data=pd.concat([sens_data,qoi_tmp])

sens_data = sens_data.reset_index(drop=True)

#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }


#multiple barplots variant by using catplot function from seaborn library
# https://seaborn.pydata.org/generated/seaborn.catplot.html

g = sns.catplot(data=sens_data,
                y="parameter", x="Gini importance",col="qoi",
                color="parameter",palette=palette,
                kind="bar", orient="h",
                col_wrap=3,height=3, aspect=1.4,
                sharex=False,sharey=False)

#set title for each subplot
g.set_titles("{col_var}: {col_name} ")

#set limits for X-axis
g.set(xlim=(0, 1))
#remove top and right spines from plots
g.despine(top=False, right=False)

#customize legend
legend_elements = [Line2D([0], [0], color='gray',linestyle="--", label='dummy'),]
for QoI, ax in g.axes_dict.items():
    ax.grid(False)
    ax.set(xlabel=r'Importance measure', ylabel='Parameter')
    ax.invert_yaxis()
    
#adjust subplots
plt.subplots_adjust(hspace=0.4, wspace=0.5)

#add text to the custom position
g.fig.suptitle("Gini importance measure by QoI", fontweight="bold",x=0.85,y=0.25,fontsize=13)

#save figure with desired name, format, dpi. 
#use f'{<variable_name>}' for filename customization using previously defined variables
g.savefig(f"4_{filename_prefix}_multi_bars_gini.png", dpi=600)