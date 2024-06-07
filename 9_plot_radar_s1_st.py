###################
# radar chart

# -*- coding: utf-8 -*-
#import required libraries
import os
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from math import pi

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'

filename='snl_no_graph_s1_st_dakota_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    
#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]


#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

##~Okabe&Ito palette
palette= ['#F0E442',          '#E69F00',
          '#D55E00',          '#0072B2',
          '#009E73',          '#56B4E9',
          '#CC79A7',          '000000',]

#create figure, set dimensions
fig, ax = plt.subplots(1,1,figsize=(8,6),subplot_kw=dict(projection="polar"))

#Create background
 
#count parameters
parameters=list(sens_data_qoi['parameter'])
num_par=len(parameters)
 
# calculate angles of each axis in the plot (divide the plot by number of parameters)
angles = [n / float(num_par) * 2 * pi for n in range(num_par)]
angles += angles[:1]
 
 
# position of the first axis :
ax.set_theta_offset(pi / 2)
ax.set_theta_direction(-1)
 
# Draw one axe per parameter + add labels

# Draw ylabels
ax.set_rlabel_position(0)
plt.ylim(0,0.7)
ticks=[0.1,0.2,0.3,0.4,0.5,0.6,0.7]
ticklabels=[f"{tick}" for tick in ticks]
plt.yticks(ticks, ticklabels, color="grey", size=7,zorder=100)

# Plot each line = each compared set of the data

#S1
measure='S1'
measure_color=palette[3]
si_values=list(sens_data_qoi[measure])
si_values += si_values[:1]

ax.plot(angles, si_values, linewidth=1, linestyle='solid', color=measure_color,label=f"${measure.replace('S','S_')}$")
ax.fill(angles, si_values, measure_color, alpha=0.1)


#ST
measure='ST'
measure_color=palette[4]
si_values=list(sens_data_qoi[measure])
si_values += si_values[:1]

ax.plot(angles, si_values, linewidth=1, linestyle='solid',color=measure_color, label=f"${measure.replace('S','S_')}$")
ax.fill(angles, si_values, measure_color, alpha=0.1)

plt.xticks(angles[:-1], parameters)
ax.tick_params(axis='x',pad=15)

# Add legend
plt.legend(loc='upper right', bbox_to_anchor=(-0.05, 0.1), fontsize=12)
fig.subplots_adjust(top=0.85,bottom=0.15,left=0.15,right=0.85)
fig.suptitle(f"Main and total Sobol' indices. QoI: {qoi}",fontsize=13,fontweight='bold',y=0.98)


#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables                                        
fig.savefig(f'9_{filename_prefix}_{qoi}_radar_s1_st.png',dpi=600)