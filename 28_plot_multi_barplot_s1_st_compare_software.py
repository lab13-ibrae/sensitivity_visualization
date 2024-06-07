# -*- coding: utf-8 -*-
#import required libraries
import os
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from matplotlib.lines import Line2D

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into one dataframe
data_folder='./sensitivity_results_datasets'
filename1='snl_no_graph_s1_st_dakota_by_qoi.xlsx'
filename_prefix=filename1.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename1))
sens_data=pd.DataFrame()
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    qoi_tmp=pd.DataFrame(data, columns=cols)
    qoi_tmp['qoi']=qoi
    qoi_tmp['software']='DAKOTA'
    sens_data=pd.concat([sens_data,qoi_tmp])

filename2='snl_no_graph_s1_st_chaospy_by_qoi.xlsx'
wb = load_workbook(filename = os.path.join(data_folder,filename2))

for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    qoi_tmp=pd.DataFrame(data, columns=cols)
    qoi_tmp['qoi']=qoi
    qoi_tmp['software']='Chaospy'
    sens_data=pd.concat([sens_data,qoi_tmp])

sens_data = sens_data.reset_index(drop=True)


sens_data_melted = sens_data.melt(id_vars=["parameter","qoi","software"], var_name='measure', value_name="value")


palette= ['0.6','0.2']


#multiple barplots variant by using catplot function from seaborn library
# https://seaborn.pydata.org/generated/seaborn.catplot.html

g = sns.catplot(data=sens_data_melted,
                #facet_kws={'sharex': False, 'sharey': False,},
                y="parameter", x="value",
                col="software",
                row="qoi",
                #color="parameter",
                palette=palette,
                hue='measure',
                kind="bar", orient="h",
                height=3, aspect=1.4,
                sharex= False,
                #margin_titles=True,
)

#set title for each subplot
g.set_titles(col_template="{col_name}",row_template="{row_name}",  fontsize=10)



#set limits for X-axis
g.set(xlim=(0, 1))
#remove top and right spines from plots
g.despine(top=False, right=False)


for QoI, ax in g.axes_dict.items():
#     ax.grid(False)
    #ax.set(xlabel=r"$S_T$", ylabel='Parameter')
    ax.set(xlabel="", ylabel='Parameter')
    ax.invert_yaxis()
g.fig.set_size_inches(8,12)


g._legend.remove()
legend_dict=g._legend_data
legend_dict={k.replace('S', r'$S_') +'$': v for k, v in legend_dict.items()}
handles = legend_dict.values()
labels = legend_dict.keys()

g.fig.legend(handles=handles, labels=labels, loc='upper center', ncol=2,bbox_to_anchor=[0.5, 0.975])

#add text to the custom position
g.fig.suptitle("Sobol' indices: software comparison", fontweight="bold",x=0.5,y=0.99,fontsize=13)

#adjust subplots
g.fig.subplots_adjust(hspace=0.4, wspace=0.5,top=0.92)

#g.fig.tight_layout()

#save figure with desired name, format, dpi. 
#use f'{<variable_name>}' for filename customization using previously defined variables
g.savefig(f"28_{filename_prefix}_multi_bars_sobol_software.png", dpi=600)




