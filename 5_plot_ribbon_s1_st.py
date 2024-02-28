# -*- coding: utf-8 -*-
import os
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)

filename='snl_no_graph_s1_st_dakota_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")
data_folder='./sensitivity_results_datasets'
wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    

#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]



#create figure, set dimensions
fig, ax = plt.subplots(1,1,figsize=(8,6))


#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }



#map palette to the parameters
sens_data_qoi['color']=sens_data_qoi['parameter'].map(palette)

#dataframe with main and total indices and renamed (formatted) dataframe columns 
sens_data_qoi_ind_only=sens_data_qoi[['parameter',
                                  'S1','ST']].set_index('parameter').rename(columns={'S1': '$S_1$',
                                                                                     'ST': '$S_T$'})


#prepare and draw barplot
#variable for bottoms
bottom = np.zeros(len(sens_data_qoi_ind_only.columns))

#loop through the parameters
for i, parameter in enumerate(sens_data_qoi_ind_only.index):

  ax.bar(sens_data_qoi_ind_only.columns, 
         sens_data_qoi_ind_only.loc[parameter], 
         bottom=bottom, label=parameter,width=0.2,
         color=palette[parameter], 

         )
  #bottom of next part is set as the top of the previous
  bottom += np.array(sens_data_qoi_ind_only.loc[parameter])

#set limits for Y-axis
ax.set_ylim(0,1.2)
#labels on the X-axis from column names (main indices and total indices)
ax.set_xticks(list(sens_data_qoi_ind_only.columns))

#extend plot's borders
box = ax.get_position()
ax.set_position([box.x0, box.y0 - box.height * 0.02,
                 box.width, box.height * 1.02])

#customize legend's labels and text
labels_legend = sens_data_qoi['parameter']+'($S_1$ = '+sens_data_qoi['S1'].map('{:,.2f}'.format)+', $S_T$ = '+sens_data_qoi['ST'].map('{:,.2f}'.format)+')'
#customize legend
plt.legend(labels_legend, loc="lower center",bbox_to_anchor=(0.5,0.15),ncol=1, title='Parameters')
handles, labels = ax.get_legend_handles_labels()
ax.legend(handles[::-1], labels[::-1])

#plot title
plt.title(f"QoI: {qoi}. Main and total Sobol' indices", y=1, fontweight="bold")

#adjust the padding
plt.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables
plt.savefig(f'5_{filename_prefix}_{qoi}_ribbon_s1_st.png',dpi=600)


