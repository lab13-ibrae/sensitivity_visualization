# -*- coding: utf-8 -*-
#import required libraries
import os
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'
filename='snl_no_graph_s1_st_dakota_by_qoi.xlsx'
filename_prefix=filename.replace(".xlsx","")
wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    
#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"

#dataframe with selected qoi
si_qoi=sens_data[qoi]

#create figure, set dimensions
fig = plt.figure(figsize=(8,6))


#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }
#map palette to the parameters
si_qoi['color']=si_qoi['parameter'].map(palette)

#prepare pie chart
#labels for the pie's sectors (only for parameters with S1>0.03)
labels = np.where(si_qoi['S1']>0.03, si_qoi['parameter'], '')
#colors
colors=si_qoi['color']

#labels for the legend with formatting
labels_legend = si_qoi['parameter']+' ($S_1$ = '+si_qoi['S1'].map('{:,.3f}'.format)+')'

#sizes of the sectors
sizes = si_qoi['S1']

#display percents only if >3
def autopct(pct):
    return ('%.1f' % pct+'%') if pct >3 else ''

#draw pie chart
_, _, autotexts=plt.pie(sizes,
                        colors=colors,
                        labels=labels,
                        #!!!!!!!!!!!!!!!!!!!!!!!!!!!
                        normalize=False,
                        startangle=90,
                        autopct=autopct,
                        radius=1,
                        labeldistance=1.05,
                        pctdistance=0.7,
                        wedgeprops={"edgecolor":"white",'linewidth': 0.5, 'antialiased': True})
#change color of percents
for autotext in autotexts:
    autotext.set_color('white')
    
#plot legend
plt.legend(labels_legend, loc="lower center",bbox_to_anchor=(0.5,0.95),ncol=2 )

#plot title
plt.title(f"Sobol' main indices ($S_1$). QoI: {qoi}\n Empty sector: fraction of variance unexplained", y=-0.05, fontweight="bold")

#adjust the padding
plt.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables                                        
plt.savefig(f'7_{filename_prefix}_{qoi}_pie.png',dpi=600)