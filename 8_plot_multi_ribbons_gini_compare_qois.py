# -*- coding: utf-8 -*-
import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")



#load results from xlsx file into one dataframe
data_folder='./sensitivity_results_datasets'
filename="snl_no_graph_gini_by_qoi.xlsx"
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data=pd.DataFrame()
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    qoi_tmp=pd.DataFrame(data, columns=cols)
    qoi_tmp['qoi']=qoi
    sens_data=pd.concat([sens_data,qoi_tmp])

sens_data = sens_data.reset_index(drop=True)

#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_pivot=sens_data.pivot_table(index='parameter',columns='qoi',values='Gini importance',sort=False)

#create figure, set dimensions
fig, ax = plt.subplots(figsize=(13,5))

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }

#prepare barplots
#variable for bottoms
bottom = np.zeros(len(sens_data_pivot.columns))
#loop through the parameters
for i, parameter in enumerate(sens_data_pivot.index):
  ax.bar(sens_data_pivot.columns, sens_data_pivot.loc[parameter], bottom=bottom, label=parameter,width=0.5,
         color=palette[parameter], 

         )
  #bottom of next part is set as the top of the previous
  bottom += np.array(sens_data_pivot.loc[parameter])

#labels on the X-axis from column names (realization numbers)
ax.set_xticks(list(sens_data_pivot.columns))

#extend plot's borders
box = ax.get_position()
ax.set_position([box.x0, box.y0 - box.height * 0.01,
                 box.width, box.height * 1.01])

#legend variants
handles, labels = ax.get_legend_handles_labels()
#legend at the bottom, parameters line by line
# def flip(items, ncol):
#     return itertools.chain(*[items[i::ncol] for i in range(ncol)])
# ax.legend(flip(handles, 3), flip(labels, 3),loc='upper center', bbox_to_anchor=(0.5, -0.1), ncol=3)
#legend at the right, parameters from the bottom to the top
ax.legend(handles[::-1], labels[::-1],loc='lower right', bbox_to_anchor=(1.2, 0.1), ncol=1,title='Parameters')
ax.tick_params(axis='x', which='major', labelsize=9.5, rotation=0)
#plot title (QoI name as is)
#ax.set_title(f"Importance measure by DFN realization number.\n QoI: {qoi}", fontweight='bold')



ax.set_title(f"Gini importance measures by QoI", fontweight='bold')

#adjust the padding
fig.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables
plt.savefig(f'8_{filename_prefix}_multi_ribbon_st.png',dpi=600)