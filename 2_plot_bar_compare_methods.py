# -*- coding: utf-8 -*-
#import required libraries
import os
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into dictionary of dataframes (one dataframe for each xlsx sheet/qoi)
data_folder='./sensitivity_results_datasets'
filename='snl_no_graph_gini_shap_qoi1.xlsx'
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data={}
for qoi in wb.sheetnames:
    ws = wb[qoi]
    data = ws.values
    cols = next(data)[0:]
    sens_data[qoi]=pd.DataFrame(data, columns=cols)
    
#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"


#dataframe with selected qoi
sens_data_qoi=sens_data[qoi]

#create figure, set dimensions
fig, ax = plt.subplots(1,1,figsize=(8,6))


#prepare chart elements
#lables by parameter names
labels = sens_data_qoi['parameter']
#positions of the bars
x_pos=np.arange(len(labels))*2

#plot bars for method 1
method1='Gini'
bar1=ax.bar(x_pos-0.4, sens_data_qoi[method1], 
            #yerr=sens_data_qoi['S1_conf'],capsize=10,
            #align='center', alpha=0.5, color='dimgray',ecolor='slategray',hatch='///')
            align='center', alpha=0.5, color='0.95',edgecolor='0.1',hatch='+++')
#plot bars for method 2
method2='SHAP'
bar2=ax.bar(x_pos+0.4, sens_data_qoi[method2],
            #yerr=sens_data_qoi['ST_conf'], capsize=10, 
            #align='center', alpha=0.5, color='dimgray',ecolor='slategray',hatch='')
            align='center', alpha=0.5, color='0.95',edgecolor='0.1',hatch='XXX')

#set Y-axis name
ax.set_ylabel('Importance measures')
#set X-axis name
ax.set_xlabel('Parameter')

#Parameter names at the X-axis
ax.set_xticks(x_pos)
ax.set_xticklabels(labels)

#turn on grid
ax.yaxis.grid(True)

#plot title
plt.title(f"Importance measures. QoI: {qoi}", y=1, fontweight="bold")

#customize plot legend
ax.legend( (bar1, bar2), (f'{method1}', f'{method2}'),handleheight=1.5 )

#adjust the padding
plt.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables                                        
plt.savefig(f'2_{filename_prefix}_{qoi}_bar_compare_methods.png',dpi=600)