# -*- coding: utf-8 -*-
import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#load results from xlsx file into one dataframe
data_folder='./sensitivity_results_datasets'
filename="snl_no_graph_by_realizations_qoi_dakota_s1_st.xlsx"
filename_prefix=filename.replace(".xlsx","")

wb = load_workbook(filename = os.path.join(data_folder,filename))
sens_data=pd.DataFrame()
for realization in wb.sheetnames:
    ws = wb[realization]
    data = ws.values
    cols = next(data)[0:]
    realization_tmp=pd.DataFrame(data, columns=cols)
    realization_tmp['realization']=realization
    sens_data=pd.concat([sens_data,realization_tmp])

sens_data = sens_data.reset_index(drop=True)

#select qoi
qoi='Peak_TotalI129_M'
#qoi="FractionofSpikeinRepository_1My"
#qoi="FractionalMassFluxfromRepo_1Myr"
#qoi="AqEb_RockEb_1Myr"
#qoi="RockAq_RockEb_1Myr"

#dataframe with selected qoi
sens_data_qoi=sens_data[sens_data['qoi']==qoi].pivot_table(index='parameter',columns='realization',values='ST',sort=False)

#create figure, set dimensions
fig, ax = plt.subplots(figsize=(10,5))

#set plot style: https://www.python-graph-gallery.com/104-seaborn-themes
sns.set_theme(style="whitegrid")

#set custom palette
# palette= {'STT':'#458E85','Intersections':'#8E1843','aveDegree':'#EACE3C',
#           'pBuffer':'#07482C','meanWPrate':'#19675D','stdWPrate':'#5B3A73',
#                       'IRF':'#AA6CAD','rateUNF':'#B51717','kGlacial':'#C14912',
#                       'permDRZ':'#C79A03','permBuffer':'#A4B16F',
#                       'dummy':'#A9A9A9'}

##~Okabe&Ito palette
palette= {
          'pBuffer':'#F0E442','meanWPrate':'#E69F00','stdWPrate':'#D55E00',
                      'IRF':'#0072B2','rateUNF':'#009E73','kGlacial':'#56B4E9',
                      'permDRZ':'#CC79A7','permBuffer':'#000000',
                      }

#prepare barplots
#variable for bottoms
bottom = np.zeros(len(sens_data_qoi.columns))
#loop through the parameters
for i, parameter in enumerate(sens_data_qoi.index):
  ax.bar(sens_data_qoi.columns, sens_data_qoi.loc[parameter], bottom=bottom, label=parameter,width=0.5,
         color=palette[parameter], 

         )
  #bottom of next part is set as the top of the previous
  bottom += np.array(sens_data_qoi.loc[parameter])

#labels on the X-axis from column names (realization numbers)
ax.set_xticks(list(sens_data_qoi.columns))

#extend plot's borders
box = ax.get_position()
ax.set_position([box.x0, box.y0 - box.height * 0.01,
                 box.width, box.height * 1.01])

#legend variants
handles, labels = ax.get_legend_handles_labels()
#legend at the bottom, parameters line by line
# def flip(items, ncol):
#     return itertools.chain(*[items[i::ncol] for i in range(ncol)])
# ax.legend(flip(handles, 3), flip(labels, 3),loc='upper center', bbox_to_anchor=(0.5, -0.1), ncol=3)
#legend at the right, parameters from the bottom to the top
ax.legend(handles[::-1], labels[::-1],loc='lower right', bbox_to_anchor=(1.2, 0.1), ncol=1,title='Parameters')

#plot title (QoI name as is)
#ax.set_title(f"Importance measure by DFN realization number.\n QoI: {qoi}", fontweight='bold')

# #dictionary for nice QoI names
# qoi_dict={"Peak_TotalI129_M": "Peak $^{129}$I concentration in aquifer (M)",
#           "FractionofSpikeinRepository_1Myr": "Fraction of Spike in repository at 1 million years",
#           "FractionalMassFluxfromRepo_1Myr":" Fractional Mass Flux of Tracer from Repository at 1 million years (1/yr)",
#           "AqEb_RockEb_1Myr":" Aquifer to East / Rock to East at 1 million years",
#           "RockAq_RockEb_1Myr":" Rock to Aquifer / Rock to East at 1 million years"}

#plot title (nice QoI name)
ax.set_title(f"Sobol' total indices by DFN realization number.\n QoI: {qoi}", fontweight='bold')

#adjust the padding
fig.tight_layout()

#save figure with desired name, format, dpi. f'{<variable_name>}' for filename customization using previously defined variables
plt.savefig(f'8_{filename_prefix}_{qoi}_multi_ribbon_st.png',dpi=600)